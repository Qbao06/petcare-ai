import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

HEADERS = [
    "timestamp", "species", "age_months",
    "symptoms", "signs_json",
    "top1", "conf1", "top2", "conf2", "top3", "conf3",
    "alert", "recommendation", "image_name",
]

class ExcelLockedError(Exception):
    pass

def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

def append_log_row(path: str, row: dict):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "LOG"
        ws.append(HEADERS)
        try:
            wb.save(path)
        except PermissionError as e:
            raise ExcelLockedError(f"Excel file is locked: {path}") from e

    wb = load_workbook(path)
    ws = wb["LOG"]
    ws.append([row.get(h, "") for h in HEADERS])
    _autosize(ws)

    try:
        wb.save(path)
    except PermissionError as e:
        raise ExcelLockedError(f"Excel file is locked: {path}") from e